"""
template_detector.py  v2
------------------------
Detect template bằng sheet name trước, fallback column match.
Đọc columns từ TemplateDef v2.
"""

from pathlib import Path
import openpyxl
from engine.config_loader import load_config, list_templates


def detect_template(file_path: str | Path) -> dict:
    path        = Path(file_path)
    wb          = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    templates = list_templates()

    # ── Strategy 1: Sheet name match ─────────────────────────────────────────
    sheet_lower = {s.lower().replace(" ","_"): s for s in sheet_names}
    for tmpl in templates:
        tmpl_key = tmpl.lower().replace(" ","_")
        if tmpl_key in sheet_lower:
            return {
                "template":   tmpl,
                "confidence": 1.0,
                "sheet_name": sheet_lower[tmpl_key],
                "matched":    [],
                "missing":    [],
                "all_scores": {t: (1.0 if t == tmpl else 0.0) for t in templates},
                "method":     "sheet_name_match",
            }

    # ── Strategy 2: Column match ──────────────────────────────────────────────
    best_score    = 0.0
    best_template = templates[0] if templates else "HR"
    best_sheet    = sheet_names[0]
    best_details  = {"matched": [], "missing": [], "extra": []}
    all_scores    = {t: 0.0 for t in templates}

    wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in sheet_names:
        ws   = wb2[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))

        for header_row in rows:
            if not header_row:
                continue
            file_cols = {str(c).strip().lower() for c in header_row if c is not None and str(c).strip()}
            if not file_cols:
                continue

            for tmpl in templates:
                tmpl_def      = load_config(tmpl)
                config_cols   = {c.name.lower() for c in tmpl_def.columns}
                required_cols = {c.name.lower() for c in tmpl_def.columns if c.is_required}

                matched  = file_cols & config_cols
                missing  = config_cols - file_cols
                extra    = file_cols - config_cols

                req_matched = matched & required_cols
                req_total   = len(required_cols) if required_cols else 1

                score = 0.7 * (len(req_matched)/req_total) + \
                        0.3 * (len(matched)/len(config_cols) if config_cols else 0)
                if file_cols and len(extra)/len(file_cols) > 0.5:
                    score *= 0.8

                if score > best_score:
                    best_score    = round(score, 4)
                    best_template = tmpl
                    best_sheet    = sheet_name
                    best_details  = {
                        "matched": sorted(matched),
                        "missing": sorted(missing),
                        "extra":   sorted(extra),
                    }
                all_scores[tmpl] = max(all_scores[tmpl], round(score, 4))
    wb2.close()

    return {
        "template":   best_template,
        "confidence": best_score,
        "sheet_name": best_sheet,
        "matched":    best_details["matched"],
        "missing":    best_details["missing"],
        "extra":      best_details["extra"],
        "all_scores": all_scores,
        "method":     "column_match",
    }