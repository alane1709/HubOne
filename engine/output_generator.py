"""
output_generator.py
-------------------
Step 5: Xuất file Excel lỗi (highlight) + JSON summary
"""

import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime


# ── Colors ────────────────────────────────────────────────────────────────────
RED_FILL    = PatternFill("solid", fgColor="FF4444")
YELLOW_FILL = PatternFill("solid", fgColor="FFD700")
GREEN_FILL  = PatternFill("solid", fgColor="00C851")
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FONT    = Font(color="CC0000", bold=True)


def generate_output(file_path: str | Path, result: dict, output_dir: str | Path) -> dict:
    """
    Tạo:
      1. Excel file: data gốc + highlight lỗi + cột error_message
      2. JSON summary

    Returns: {"excel_path": ..., "summary_path": ..., "summary": ...}
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem       = Path(file_path).stem
    excel_out  = output_dir / f"{stem}_errors_{timestamp}.xlsx"
    summary_out = output_dir / f"{stem}_summary_{timestamp}.json"

    # Build error lookup: {row_id: [errors]}
    error_map: dict[int, list[dict]] = {}
    for e in result["errors"]:
        error_map.setdefault(e["row_id"], []).append(e)

    # Column-level errors (row_id == 0 = column-level)
    col_errors: dict[str, list[dict]] = {}
    for e in result["errors"]:
        if e["row_id"] == 0:
            col_errors.setdefault(e["column_name"].lower(), []).append(e)

    # ── Read original file ────────────────────────────────────────────────────
    src_wb = openpyxl.load_workbook(file_path)
    src_ws = src_wb.active

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Validation Result"

    # ── Copy header row ───────────────────────────────────────────────────────
    header_row = [cell.value for cell in src_ws[1]]
    header_lower = [str(h).strip().lower() if h else "" for h in header_row]

    for col_idx, val in enumerate(header_row, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Add error_message column header
    err_col = len(header_row) + 1
    err_header = out_ws.cell(row=1, column=err_col, value="⚠ error_message")
    err_header.fill = PatternFill("solid", fgColor="8B0000")
    err_header.font = HEADER_FONT
    err_header.alignment = Alignment(horizontal="center")

    # ── Copy data rows + highlight ────────────────────────────────────────────
    for row_idx, src_row in enumerate(src_ws.iter_rows(min_row=2, values_only=True), start=2):
        row_errors = error_map.get(row_idx, [])
        row_err_cols = {e["column_name"].lower() for e in row_errors}
        row_messages = []

        for col_idx, val in enumerate(src_row, 1):
            cell = out_ws.cell(row=row_idx, column=col_idx, value=val)
            col_name = header_lower[col_idx - 1] if col_idx - 1 < len(header_lower) else ""

            # Highlight cell nếu có lỗi
            if col_name in row_err_cols:
                matching = [e for e in row_errors if e["column_name"].lower() == col_name]
                severity = matching[0]["severity"] if matching else "warning"
                cell.fill = RED_FILL if severity == "critical" else YELLOW_FILL
                for e in matching:
                    row_messages.append(f"[{e['error_type']}] {e['message']}")

        # error_message column
        msg_cell = out_ws.cell(row=row_idx, column=err_col,
                               value=" | ".join(row_messages) if row_messages else "✓ OK")
        if row_messages:
            msg_cell.font = RED_FONT
        else:
            msg_cell.fill = PatternFill("solid", fgColor="E8F5E9")

    # ── Auto-width columns ────────────────────────────────────────────────────
    for col in out_ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        out_ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # ── Sheet 2: Error Summary ────────────────────────────────────────────────
    sum_ws = out_wb.create_sheet("Summary")

    summary_rows = [
        ["Field", "Value"],
        ["Template",      result["template"]],
        ["Total Rows",    result["total_rows"]],
        ["Error Rows",    result["error_rows"]],
        ["Error %",       f"{result['error_pct']}%"],
        ["Critical",      result["critical"]],
        ["Warnings",      result["warnings"]],
        ["Status",        result["status"].upper()],
        ["Generated At",  datetime.now().isoformat()],
    ]

    for r_idx, row in enumerate(summary_rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = sum_ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            elif c_idx == 1:
                cell.font = Font(bold=True)

    # Status color
    status_cell = sum_ws.cell(row=8, column=2)
    if result["status"] == "fail":
        status_cell.fill = RED_FILL
        status_cell.font = Font(bold=True, color="FFFFFF")
    elif result["status"] == "warning":
        status_cell.fill = YELLOW_FILL
    else:
        status_cell.fill = GREEN_FILL
        status_cell.font = Font(bold=True, color="FFFFFF")

    sum_ws.column_dimensions["A"].width = 20
    sum_ws.column_dimensions["B"].width = 30

    # ── Save Excel ────────────────────────────────────────────────────────────
    out_wb.save(excel_out)

    # ── Save JSON summary ─────────────────────────────────────────────────────
    summary = {
        "template":     result["template"],
        "total_rows":   result["total_rows"],
        "error_rows":   result["error_rows"],
        "error_pct":    result["error_pct"],
        "critical":     result["critical"],
        "warnings":     result["warnings"],
        "status":       result["status"],
        "generated_at": datetime.now().isoformat(),
        "errors":       result["errors"],
    }

    with open(summary_out, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    return {
        "excel_path":   str(excel_out),
        "summary_path": str(summary_out),
        "summary":      summary,
    }


def generate_output_all(file_path: str | Path, results: dict, output_dir: str | Path) -> dict:
    """
    Tạo output cho /validate-all (nhiều template):
      1. Excel multi-sheet: mỗi sheet = kết quả 1 template
      2. JSON summary tổng hợp

    results: {template_name: validation_result}
    Returns: {"excel_path": ..., "summary_path": ...}
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem        = Path(file_path).stem
    excel_out   = output_dir / f"{stem}_multi_{timestamp}.xlsx"
    summary_out = output_dir / f"{stem}_multi_summary_{timestamp}.json"

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)  # remove default sheet

    all_errors = []

    for tmpl_name, result in results.items():
        sheet_title = tmpl_name[:31].replace("/", "-").replace("\\", "-")
        ws = out_wb.create_sheet(sheet_title)

        # Build error lookup
        error_map: dict[int, list[dict]] = {}
        for e in result["errors"]:
            error_map.setdefault(e["row_id"], []).append(e)

        # Read source sheet
        src_wb = openpyxl.load_workbook(file_path)
        src_ws = src_wb[result.get("sheet_name", 0)] if result.get("sheet_name") else src_wb.active
        src_wb.close()

        # Header row
        header_row = [cell.value for cell in src_ws[1]]
        for col_idx, val in enumerate(header_row, 1):
            cell = ws.cell(row=1, column=col_idx, value=val)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Add error_message column
        err_col = len(header_row) + 1
        err_header = ws.cell(row=1, column=err_col, value="⚠ error_message")
        err_header.fill = PatternFill("solid", fgColor="8B0000")
        err_header.font = HEADER_FONT

        # Copy data + highlight
        for row_idx, src_row in enumerate(src_ws.iter_rows(min_row=2, values_only=True), start=2):
            row_errors = error_map.get(row_idx, [])
            row_err_cols = {e["column_name"].lower() for e in row_errors}
            row_messages = []

            for col_idx, val in enumerate(src_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                header_lower = [str(h).strip().lower() if h else "" for h in header_row]
                col_name = header_lower[col_idx - 1] if col_idx - 1 < len(header_lower) else ""

                if col_name in row_err_cols:
                    matching = [e for e in row_errors if e["column_name"].lower() == col_name]
                    severity = matching[0]["severity"] if matching else "warning"
                    cell.fill = RED_FILL if severity == "critical" else YELLOW_FILL
                    for e in matching:
                        row_messages.append(f"[{e['error_type']}] {e['message']}")

            msg_cell = ws.cell(row=row_idx, column=err_col,
                               value=" | ".join(row_messages) if row_messages else "✓ OK")
            if row_messages:
                msg_cell.font = RED_FONT
            else:
                msg_cell.fill = PatternFill("solid", fgColor="E8F5E9")

        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Sheet 2: Summary per template
        sum_ws = out_wb.create_sheet(f"{sheet_title}_Summary")
        sum_rows = [
            ["Field", "Value"],
            ["Template",      result["template"]],
            ["Total Rows",    result["total_rows"]],
            ["Error Rows",    result["error_rows"]],
            ["Error %",       f"{result['error_pct']}%"],
            ["Critical",      result["critical"]],
            ["Warnings",      result["warnings"]],
            ["Status",        result["status"].upper()],
        ]
        for r_idx, row in enumerate(sum_rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = sum_ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                elif c_idx == 1:
                    cell.font = Font(bold=True)
        sum_ws.column_dimensions["A"].width = 20
        sum_ws.column_dimensions["B"].width = 30

        all_errors.extend(result.get("errors", []))

    out_wb.save(excel_out)

    # JSON summary
    summary = {
        "total_templates": len(results),
        "generated_at": datetime.now().isoformat(),
        "sheets": {
            tmpl: {
                "total_rows": r["total_rows"],
                "error_rows": r["error_rows"],
                "error_pct":  r["error_pct"],
                "critical":   r["critical"],
                "warnings":    r["warnings"],
                "status":      r["status"],
            }
            for tmpl, r in results.items()
        },
        "errors": all_errors,
    }

    with open(summary_out, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    return {
        "excel_path":   str(excel_out),
        "summary_path": str(summary_out),
    }
