"""
create_sample_data.py
---------------------
Generates sample Excel files for testing the upload API.

Run:
    python create_sample_data.py
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from openpyxl.styles import PatternFill, Font

def create_hr_sample():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HR Data"

    # Header row
    headers = ["employee_id", "full_name", "email", "phone", "department", "start_date", "end_date", "salary"]
    header_fill = PatternFill("solid", fgColor="2196F3")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Sample data rows (mix of valid + intentional errors for testing)
    rows = [
        ["EMP001", "Nguyen Van A", "nguyenvana@company.com", "0901234567", "Engineering", "2023-01-15", None,       15000000],
        ["EMP002", "Tran Thi B",  "tranthib@company.com",   "0912345678", "HR",          "2022-06-01", None,       12000000],
        ["EMP003", "Le Van C",    "invalid-email",           None,         "Finance",     "2023-03-10", None,       18000000],  # bad email
        ["EMP001", "Duplicate",   "dup@company.com",         "0934567890", "IT",          "2021-09-01", None,       20000000],  # duplicate ID
        ["EMP005", None,          "emp5@company.com",        "0956789012", "Marketing",   "2023-07-20", None,       11000000],  # null full_name
        ["EMP006", "Pham Thi F",  "phamf@company.com",       "0967890123", "Engineering", "2024-01-01", "2023-12-31", 16000000], # end_date < start_date
        ["EMP007", "Hoang Van G", "hoangg@company.com",      "0978901234", "Sales",       "2022-11-15", None,       -5000],    # negative salary
    ]

    for row in rows:
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    path = Path(__file__).parent / "uploads" / "sample_HR.xlsx"
    path.parent.mkdir(exist_ok=True)
    wb.save(path)
    print(f"✅ Sample HR file created: {path}")
    return path


if __name__ == "__main__":
    create_hr_sample()
    print("\nSample files ready in uploads/ folder.")
    print("Now start the API: uvicorn api.main:app --reload")
